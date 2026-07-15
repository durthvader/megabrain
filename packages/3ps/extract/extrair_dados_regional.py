"""Extrai os indicadores do projeto 3Ps para qualquer regional a partir dos exports de BI.

Generaliza o extrator específico do Ceará para regionais com N clusters
(nao apenas 2). Os clusters sao detectados automaticamente a partir da linha
"R.. | ..." no arquivo 'Produtividade B2C  Visão Daily 3P's  Diário D-1.xlsx' --
nao precisam ser digitados a mao.

Os nomes de arquivo variam um pouco entre pastas (o Chrome adiciona " (1)",
" (2)" em downloads repetidos) -- por isso a busca de arquivo aqui e por
prefixo normalizado, nao por nome exato.

O script NAO escreve titulos, decisoes ou linhas de FCA: isso continua
exigindo leitura humana dos numeros e fica no script build_3ps_presentation_<regiao>.py.

Uso:
    python packages/3ps/extract/extrair_dados_regional.py --pasta "C:/Users/rogerio.fonseca/Downloads/comgsp" \
        --regional "R5 | DF, MG, SP INT, GO E TO" --saida "projects/3ps-comgsp/data/processed/dados.json"
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]

META_POR_FAIXA = {
    "1. Menos de 30 dias": 3.0,
    "2. De 30 a 59 dias": 3.0,
    "3. De 60 a 89 dias": 3.5,
    "4. Mais de 90 dias": 4.0,
}
FAIXAS_NOVATO = {"1. Menos de 30 dias", "2. De 30 a 59 dias"}

METAS_PAINEL = {
    "presenca": 0.95,
    "prazo_24h": 0.85,
    "fila_dias": 1.0,
}


def normalizar(texto: str) -> str:
    texto = "".join(c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s*\(\d+\)\s*\.xlsx$", ".xlsx", texto, flags=re.IGNORECASE)
    texto = re.sub(r"\s+", " ", texto).strip().lower()
    return texto


def encontrar_arquivo(pasta: Path, prefixo: str) -> Path:
    alvo = normalizar(prefixo)
    candidatos = [p for p in pasta.glob("*.xlsx") if normalizar(p.stem + ".xlsx").startswith(alvo)]
    if not candidatos:
        raise FileNotFoundError(
            f"Nenhum arquivo encontrado com prefixo '{prefixo}' em {pasta}.\n"
            "Confirme se o export do BI equivalente foi colado na pasta."
        )
    candidatos.sort(key=lambda p: (0 if "(" not in p.stem else 1, p.stat().st_mtime), reverse=False)
    return candidatos[0]


def carregar(pasta: Path, prefixo: str):
    caminho = encontrar_arquivo(pasta, prefixo)
    return openpyxl.load_workbook(caminho, data_only=True, read_only=False)


def linhas(ws, header_row: int = 1) -> list[dict]:
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in vals):
            continue
        out.append(dict(zip(headers, vals)))
    return out


def detectar_clusters(pasta: Path, regional: str) -> dict:
    """Le a linha 'Regional' + as linhas de Cluster subsequentes no arquivo
    Visão Daily D-1 e monta {"C.xx | Nome": {"codigo": "C.xx", "nome": "Nome"}}."""
    wb = carregar(pasta, "Produtividade B2C  Visão Daily 3P's  Diário D-1")
    ws = wb["Export"]
    clusters: dict[str, dict] = {}
    dentro = False
    for r in range(2, ws.max_row + 1):
        reg = ws.cell(r, 1).value
        cluster = ws.cell(r, 2).value
        if reg:
            dentro = str(reg).strip() == regional.strip()
            continue
        if dentro and cluster and cluster != "Total" and "|" in str(cluster):
            codigo, nome = [p.strip() for p in str(cluster).split("|", 1)]
            clusters[str(cluster)] = {"codigo": codigo, "nome": nome}
    if not clusters:
        raise ValueError(
            f"Nenhum cluster encontrado para a regional '{regional}'. "
            "Confira o texto exato na coluna Regional do arquivo Visão Daily."
        )
    return clusters


def extrair_periodo(pasta: Path) -> tuple[str, str]:
    wb = carregar(pasta, "Analítico  Nominal - Periodo Filtrado")
    ws = wb["Export"]
    texto = ws.cell(2, 1).value or ""
    match = re.match(r"(\d{2}/\d{2}/\d{4})\s*a\s*(\d{2}/\d{2}/\d{4})", str(texto))
    if not match:
        raise ValueError(f"Não entendi o período filtrado: {texto!r}")
    ini, fim = match.groups()
    di, mi, ai = ini.split("/")
    df, mf, af = fim.split("/")
    return f"{ai}-{mi}-{di}", f"{af}-{mf}-{df}"


def extrair_fila(pasta: Path, clusters: dict, data_fim_iso: str) -> dict:
    from datetime import date

    from openpyxl.utils import range_boundaries

    wb = carregar(pasta, "Tecnicos B2C sem produção em Maio.26")
    if "hist Backlog" not in wb.sheetnames:
        return {}
    ws = wb["hist Backlog"]

    col_data = {}
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        if min_row <= 2 <= max_row:
            val = ws.cell(min_row, min_col).value
            if val is not None and hasattr(val, "date"):
                for c in range(min_col, max_col + 1):
                    col_data[c] = val.date()

    alvo = date.fromisoformat(data_fim_iso)
    datas_disponiveis = sorted(set(col_data.values()))
    data_usada = alvo if alvo in datas_disponiveis else max((d for d in datas_disponiveis if d <= alvo), default=None)
    if data_usada is None:
        return {}

    colunas_alvo = [c for c, d in col_data.items() if d == data_usada]
    col_backlog = next((c for c in colunas_alvo if ws.cell(3, c).value == "BACKLOG"), None)
    col_estoque = next((c for c in colunas_alvo if ws.cell(3, c).value == "DIAS ESTOQUE"), None)
    if col_backlog is None or col_estoque is None:
        return {}

    resultado = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label:
            continue
        for info in clusters.values():
            if info["codigo"] in str(label):
                resultado[info["codigo"]] = {
                    "backlog_os": ws.cell(r, col_backlog).value,
                    "fila_dias": ws.cell(r, col_estoque).value,
                }
    resultado["data_referencia"] = data_usada.isoformat()
    return resultado


def extrair_matriz_semanal(pasta: Path, prefixo: str, coluna_alvo: str, clusters: dict) -> dict:
    """Le um arquivo 'Produtividade Matriz Dinamica X' (Presença, Prazo de
    Reparo 24hrs, Produtividade Esperada, etc.) que tem colunas de semana
    (S26, S27, S28, S29...) e de dia (07/07, 08/07...) e devolve
    {codigo_cluster: valor} para a coluna pedida (ex.: 'S28' ou '10/07')."""
    wb = carregar(pasta, prefixo)
    ws = wb["Export"]
    col_alvo = None
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value) == coluna_alvo:
            col_alvo = c
            break
    if col_alvo is None:
        return {}
    resultado = {}
    for r in range(2, ws.max_row + 1):
        cluster = ws.cell(r, 2).value
        if not cluster or "|" not in str(cluster):
            continue
        info = clusters.get(str(cluster))
        if info:
            resultado[info["codigo"]] = ws.cell(r, col_alvo).value
    return resultado


def extrair_painel_diario(pasta: Path, clusters: dict) -> dict:
    wb = carregar(pasta, "Produtividade B2C  Visão Daily 3P's  Diário D-1")
    dados = {}
    for linha in linhas(wb["Export"]):
        cluster = str(linha.get("Cluster") or "")
        if cluster not in clusters:
            continue
        info = clusters[cluster]
        dados[info["codigo"]] = {
            "cluster": info["codigo"],
            "nome": info["nome"],
            "folha_meta": linha.get("Folha Meta"),
            "presenca": linha.get("Presença"),
            "producao_pct_esperado": linha.get("% Prod. Esperado"),
            "prazo_24h": linha.get("Prazo Reparos 24hrs"),
            "cumprimento_agenda": linha.get("Cumprimento de Agenda"),
        }
    faltando = set(info["codigo"] for info in clusters.values()) - set(dados)
    if faltando:
        raise ValueError(f"Clusters não encontrados no painel diário: {faltando}")
    return dados


def dias_trabalhados_por_funcid(pasta: Path) -> dict[str, float]:
    """Soma 'Dias Trab.' por FUNCID no arquivo de quartil Baremo -- que pode
    vir quebrado por semana (S27/S28/...) ou já agregado no período inteiro."""
    wb = carregar(pasta, "Analítico Nominal por Quartil de Produtividade Baremo OK dos Técnicos  Período Filtrado")
    ws = wb["Export"]
    linhas_q = linhas(ws)
    tem_semana = any("Semana Ano" in row for row in linhas_q[:1])
    dias: dict[str, float] = defaultdict(float)
    if tem_semana:
        for row in linhas_q:
            fid = row.get("FUNCID") or row.get("Funcid")
            if fid:
                dias[fid] += row.get("Dias Trab.") or 0
    else:
        for row in linhas_q:
            fid = row.get("FUNCID") or row.get("Funcid")
            if fid:
                dias[fid] = row.get("Dias Trab.") or 0
    return dias


def extrair_nominal(pasta: Path, clusters: dict) -> tuple[list[dict], dict]:
    wb_nominal = carregar(pasta, "Analítico  Nominal - Periodo Filtrado")
    nominal = [n for n in linhas(wb_nominal["Export"]) if n.get("FUNCID")]

    dias_por_id = dias_trabalhados_por_funcid(pasta)

    wb_ferias = carregar(pasta, "Analitico  Acompanhamento Técnicos")
    dias_ferias_por_id: dict[str, int] = defaultdict(int)
    for r in linhas(wb_ferias["Export"]):
        if r.get("periodo_ferias"):
            dias_ferias_por_id[r["FUNCID"]] += 1

    tecnicos = []
    for n in nominal:
        fid = n["FUNCID"]
        dias_trab = dias_por_id.get(fid, 0) or 0
        faixa = n.get("Faixa Tempo de Casa")
        meta_dia = META_POR_FAIXA.get(faixa, 4.0)
        volume_ok = n.get("Volume OK") or 0
        gap = max(meta_dia * dias_trab - volume_ok, 0)
        cluster_raw = n.get("Cluster")
        cluster_info = clusters.get(cluster_raw, {"codigo": cluster_raw, "nome": cluster_raw})
        ferias_plena = fid in dias_ferias_por_id and dias_trab == 0
        tecnicos.append({
            "funcid": fid,
            "nome": n["Técnico"],
            "cluster": cluster_info["codigo"],
            "ga": n.get("GA (Gestor de Área)"),
            "go": n.get("GO (Gestor de Operações)"),
            "faixa_tempo_casa": faixa,
            "dias_trabalhados": dias_trab,
            "volume_ok": volume_ok,
            "meta_dia": meta_dia,
            "rate": (volume_ok / dias_trab) if dias_trab else 0,
            "gap": round(gap, 2),
            "sem_baixa_periodo": bool(n.get("Sem Baixa no Periodo Filtrado")),
            "ferias_plena": ferias_plena,
        })

    excluidos = [t for t in tecnicos if t["ferias_plena"]]
    elegiveis = [t for t in tecnicos if not t["ferias_plena"]]

    resumo = {
        "total_nominal": len(tecnicos),
        "excluidos_ferias_plena": [{"nome": t["nome"], "cluster": t["cluster"], "ga": t["ga"]} for t in excluidos],
        "elegiveis": len(elegiveis),
    }
    return elegiveis, resumo


def bucketizar(elegiveis: list[dict]) -> dict:
    na_meta = [t for t in elegiveis if t["dias_trabalhados"] > 0 and t["rate"] >= t["meta_dia"]]
    sem_producao = [t for t in elegiveis if t["dias_trabalhados"] == 0]
    novato_abaixo = [
        t for t in elegiveis
        if t["dias_trabalhados"] > 0 and t["rate"] < t["meta_dia"] and t["faixa_tempo_casa"] in FAIXAS_NOVATO
    ]
    veterano_abaixo = [
        t for t in elegiveis
        if t["dias_trabalhados"] > 0 and t["rate"] < t["meta_dia"] and t["faixa_tempo_casa"] not in FAIXAS_NOVATO
    ]
    veteranos_total = len([t for t in elegiveis if t["faixa_tempo_casa"] not in FAIXAS_NOVATO])
    return {
        "na_meta": len(na_meta),
        "novato_abaixo": len(novato_abaixo),
        "veterano_abaixo": len(veterano_abaixo),
        "sem_producao": len(sem_producao),
        "sem_producao_nomes": [
            {"nome": t["nome"], "cluster": t["cluster"], "ga": t["ga"], "faixa": t["faixa_tempo_casa"]}
            for t in sem_producao
        ],
        "veteranos_total": veteranos_total,
        "veteranos_pct": round(veteranos_total / len(elegiveis) * 100, 1) if elegiveis else None,
    }


def tiers_pct_meta(elegiveis: list[dict]) -> dict:
    sem_producao = [t for t in elegiveis if t["dias_trabalhados"] == 0]
    ativos = [t for t in elegiveis if t["dias_trabalhados"] > 0]
    for t in ativos:
        t["pct_meta"] = t["rate"] / t["meta_dia"] * 100 if t["meta_dia"] else 0
    critico = [t for t in ativos if t["pct_meta"] < 70]
    medio = [t for t in ativos if 70 <= t["pct_meta"] < 100]
    na_meta = [t for t in ativos if t["pct_meta"] >= 100]
    gap_total = sum(t["gap"] for t in elegiveis)
    gap_critico = sum(t["gap"] for t in critico)
    return {
        "sem_producao": len(sem_producao),
        "critico_abaixo_70": len(critico),
        "medio_70_100": len(medio),
        "na_meta_acima_100": len(na_meta),
        "gap_total": round(gap_total, 1),
        "gap_critico_pct": round(gap_critico / gap_total * 100) if gap_total else None,
    }


def gap_por_cluster_e_ga(elegiveis: list[dict]) -> dict:
    por_cluster = defaultdict(float)
    por_ga = defaultdict(lambda: defaultdict(float))
    for t in elegiveis:
        por_cluster[t["cluster"]] += t["gap"]
        por_ga[t["cluster"]][t["ga"]] += t["gap"]
    gap_total = sum(por_cluster.values())
    top_individuos = sorted(elegiveis, key=lambda t: -t["gap"])[:8]
    gap_top8 = sum(t["gap"] for t in top_individuos)
    return {
        "total": round(gap_total, 1),
        "por_cluster": {k: round(v, 1) for k, v in por_cluster.items()},
        "por_ga": {
            cluster: sorted(
                [{"ga": ga, "gap": round(v, 1)} for ga, v in gas.items()],
                key=lambda x: -x["gap"],
            )
            for cluster, gas in por_ga.items()
        },
        "top8_individuos": [
            {"nome": t["nome"], "cluster": t["cluster"], "ga": t["ga"], "gap": t["gap"],
             "dias_trabalhados": t["dias_trabalhados"], "volume_ok": t["volume_ok"]}
            for t in top_individuos
        ],
        "top8_pct_do_total": round(gap_top8 / gap_total * 100) if gap_total else None,
    }


def referencias_a_replicar(elegiveis: list[dict], minimo_dias: int = 5, top_n: int = 8) -> list[dict]:
    candidatos = [t for t in elegiveis if t["dias_trabalhados"] >= minimo_dias]
    candidatos.sort(key=lambda t: -t["rate"])
    return [
        {"nome": t["nome"], "cluster": t["cluster"], "rate": round(t["rate"], 2), "dias_trabalhados": t["dias_trabalhados"]}
        for t in candidatos[:top_n]
    ]


def analise_ga(elegiveis: list[dict], ga_nome: str) -> dict:
    membros = [t for t in elegiveis if t["ga"] == ga_nome]
    if not membros:
        return {}
    abaixo = sum(1 for t in membros if t["dias_trabalhados"] == 0 or t["rate"] < t["meta_dia"])
    meta_total = sum(t["meta_dia"] * t["dias_trabalhados"] for t in membros)
    vol_total = sum(t["volume_ok"] for t in membros)
    return {
        "ga": ga_nome,
        "headcount": len(membros),
        "abaixo_da_meta": abaixo,
        "capacidade_pct": round(vol_total / meta_total * 100) if meta_total else None,
    }


def extrair_improdutividade(pasta: Path) -> dict:
    def pares(prefixo: str) -> list[tuple[str, int]]:
        wb = carregar(pasta, prefixo)
        ws = wb["Sheet1"]
        out = []
        for r in range(4, ws.max_row + 1):
            k, v = ws.cell(r, 1).value, ws.cell(r, 2).value
            if k is None or v is None:
                continue
            out.append((k, v))
        return out

    motivos = pares("Improdutivas  Motivos Baixa")
    ranking_ga = pares("Ranking  GA's Improdutivas")
    ranking_tecnicos = pares("Ranking  Técnicos Improdutivas")
    total_geral = sum(v for _, v in ranking_tecnicos)
    total_com_motivo = sum(v for _, v in motivos)
    top4 = motivos[:4]
    top4_pct = round(sum(v for _, v in top4) / total_com_motivo * 100) if total_com_motivo else None
    return {
        "total": total_geral,
        "com_motivo": total_com_motivo,
        "sem_motivo": total_geral - total_com_motivo,
        "motivos_top4": [{"motivo": k, "quantidade": v} for k, v in top4],
        "top4_pct_dos_classificados": top4_pct,
        "ranking_ga": [{"ga": k, "quantidade": v} for k, v in ranking_ga if k],
        "ranking_tecnicos_top10": [{"nome": k, "quantidade": v} for k, v in ranking_tecnicos[:10]],
    }


def extrair_wow(pasta: Path) -> dict:
    wb = carregar(pasta, "Produtividade Baremo OK dos Técnicos  WoW")
    ws = wb["Export"]
    s27 = ws.cell(3, 2).value
    s28 = ws.cell(3, 3).value
    variacao = (s28 - s27) / s27 if s27 else None
    return {"produtividade_s27": s27, "produtividade_s28": s28, "variacao_pct": round(variacao * 100, 1) if variacao is not None else None}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pasta", type=Path, required=True)
    parser.add_argument("--regional", type=str, required=True, help='Ex: "R5 | DF, MG, SP INT, GO E TO"')
    parser.add_argument("--saida", type=Path, required=True)
    parser.add_argument(
        "--semana-fechamento", type=str, default=None,
        help=(
            "Se a Visão Daily D-1 cair numa semana quase vazia (ex.: so 1-2 dias "
            "da nova semana), use este parametro para pegar Presença/Prazo/Produção "
            "de uma semana ja fechada nos arquivos '% Presença', '% Prazo de Reparo "
            "24hrs' e '% Produtividade B2C Esperada  OK' (coluna, ex.: 'S28')."
        ),
    )
    parser.add_argument(
        "--data-fechamento", type=str, default=None,
        help="Data (DD/MM, ex.: '10/07') usada para a Fila quando --semana-fechamento é informado.",
    )
    args = parser.parse_args()

    if not args.pasta.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {args.pasta}")

    clusters = detectar_clusters(args.pasta, args.regional)
    periodo_inicio, periodo_fim = extrair_periodo(args.pasta)
    painel = extrair_painel_diario(args.pasta, clusters)

    data_fila_alvo = periodo_fim
    if args.semana_fechamento:
        presenca_sem = extrair_matriz_semanal(args.pasta, "% Presença", args.semana_fechamento, clusters)
        prazo_sem = extrair_matriz_semanal(args.pasta, "% Prazo de Reparo 24hrs", args.semana_fechamento, clusters)
        producao_sem = extrair_matriz_semanal(args.pasta, "% Produtividade B2C Esperada  OK", args.semana_fechamento, clusters)
        for codigo, valores in painel.items():
            if codigo in presenca_sem:
                valores["presenca"] = presenca_sem[codigo]
            if codigo in prazo_sem:
                valores["prazo_24h"] = prazo_sem[codigo]
            if codigo in producao_sem:
                valores["producao_pct_esperado"] = producao_sem[codigo]
        if args.data_fechamento:
            dia, mes = args.data_fechamento.split("/")
            ano = periodo_fim.split("-")[0]
            data_fila_alvo = f"{ano}-{mes}-{dia}"

    fila = extrair_fila(args.pasta, clusters, data_fila_alvo)
    elegiveis, resumo_nominal = extrair_nominal(args.pasta, clusters)
    buckets = bucketizar(elegiveis)
    tiers = tiers_pct_meta(elegiveis)
    gap = gap_por_cluster_e_ga(elegiveis)
    referencias = referencias_a_replicar(elegiveis)
    improdutividade = extrair_improdutividade(args.pasta)
    wow = extrair_wow(args.pasta)

    for codigo, valores in painel.items():
        if codigo in fila:
            valores["fila_dias"] = fila[codigo]["fila_dias"]
            valores["fila_backlog_os"] = fila[codigo]["backlog_os"]

    ga_destaque = None
    for cluster_gas in gap["por_ga"].values():
        if cluster_gas and (ga_destaque is None or cluster_gas[0]["gap"] > ga_destaque["gap"]):
            ga_destaque = cluster_gas[0]
    detalhe_ga_destaque = analise_ga(elegiveis, ga_destaque["ga"]) if ga_destaque else None

    payload = {
        "gerado_de": str(args.pasta),
        "regional": args.regional,
        "clusters": {info["codigo"]: info["nome"] for info in clusters.values()},
        "periodo": {"inicio": periodo_inicio, "fim": periodo_fim},
        "painel_fechamento": args.semana_fechamento or "visão diária (D-1)",
        "fila_data_referencia": fila.get("data_referencia"),
        "painel": painel,
        "metas": METAS_PAINEL,
        "nominal": resumo_nominal,
        "buckets_producao": buckets,
        "tiers_pct_meta": tiers,
        "gap": gap,
        "ga_maior_gap": detalhe_ga_destaque,
        "referencias_a_replicar": referencias,
        "improdutividade": improdutividade,
        "wow_produtividade_baremo": wow,
    }

    args.saida.parent.mkdir(parents=True, exist_ok=True)
    args.saida.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK: {args.saida}")
    print(json.dumps({
        "clusters": payload["clusters"],
        "tecnicos_elegiveis": len(elegiveis),
        "gap_total": gap["total"],
        "improdutivas_total": improdutividade["total"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
