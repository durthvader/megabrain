"""Importa e consolida fontes complementares de produtividade do projeto 3Ps."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import urllib.parse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from importar_workbook_3ps import (
    SupabaseRest,
    create_base,
    json_value,
    normalize_key,
    parse_env,
    post_batches,
    raw_rows,
)


def classify(path: Path) -> str:
    name = normalize_key(path.stem)
    if "share_por_quartil" in name:
        return "share_quartil"
    if "nominal_por_quartil" in name:
        return "nominal_quartil"
    if "acompanhamento_tecnicos_segmento_b2c" in name:
        return "resumo_diario"
    if "analitico_acompanhamento_tecnicos" in name:
        return "detalhe_diario"
    if "nominal_periodo_filtrado" in name or "nominal_periodo" in name:
        return "nominal_periodo"
    raise ValueError(f"Fonte complementar nao reconhecida: {path.name}")


def safe_name(path: Path) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", path.name)


def load_clean(path: Path, funcid_position: int | None = None) -> pd.DataFrame:
    frame = pd.read_excel(path)
    if funcid_position is not None:
        frame = frame[frame.iloc[:, funcid_position].notna()].copy()
    return frame


def cluster_name(value) -> tuple[str, str]:
    text = str(value)
    if "C.26" in text:
        return "C.26", "Ceara Interior"
    if "C.27" in text:
        return "C.27", "Fortaleza"
    return "", text


def target_for_tenure(value) -> float:
    text = str(value)
    if "60 a 89" in text:
        return 3.5
    if "90" in text:
        return 4.0
    return 3.0


def value_or_none(value):
    if pd.isna(value):
        return None
    return json_value(value)


def build_curated(paths_by_role: dict[str, Path]) -> list[dict]:
    nominal = load_clean(paths_by_role["nominal_quartil"], 9)
    nominal_period = load_clean(paths_by_role["nominal_periodo"], 9)
    detail = pd.read_excel(paths_by_role["detalhe_diario"])
    daily = pd.read_excel(paths_by_role["resumo_diario"])
    daily = daily[pd.to_datetime(daily.iloc[:, 0], errors="coerce").notna()].copy()

    period_by_id = nominal_period.set_index(nominal_period.columns[9])
    output: list[dict] = []
    technician_records: list[dict] = []

    for _, row in nominal.iterrows():
        funcid = row.iloc[9]
        code, cluster = cluster_name(row.iloc[3])
        meta = target_for_tenure(row.iloc[12])
        productivity = float(row.iloc[15]) if not pd.isna(row.iloc[15]) else 0.0
        days_worked = float(row.iloc[14]) if not pd.isna(row.iloc[14]) else 0.0
        volume = float(row.iloc[13]) if not pd.isna(row.iloc[13]) else 0.0
        extra = period_by_id.loc[funcid] if funcid in period_by_id.index else None
        if isinstance(extra, pd.DataFrame):
            extra = extra.iloc[0]
        sem_baixa_value = (
            pd.to_numeric(pd.Series([extra.iloc[14]]), errors="coerce").iloc[0]
            if extra is not None else np.nan
        )
        record = {
            "tipo_registro": "tecnico_produtividade_periodo",
            "estado": "CE",
            "cluster_codigo": code,
            "cluster": cluster,
            "periodo": value_or_none(row.iloc[0]),
            "semana": value_or_none(row.iloc[1]),
            "tecnico": value_or_none(row.iloc[4]),
            "cargo": value_or_none(row.iloc[5]),
            "gerente_geral": value_or_none(row.iloc[6]),
            "gestor_operacoes": value_or_none(row.iloc[7]),
            "gestor_area": value_or_none(row.iloc[8]),
            "funcid": value_or_none(funcid),
            "id_ofs": value_or_none(row.iloc[10]),
            "admissao": value_or_none(row.iloc[11]),
            "faixa_tempo_casa": value_or_none(row.iloc[12]),
            "meta_produtividade": meta,
            "volume_produtivo": volume,
            "dias_trabalhados": days_worked,
            "produtividade": productivity,
            "gap_meta": productivity - meta,
            "atingiu_meta": productivity >= meta,
            "gap_os_meta": max(meta * days_worked - volume, 0.0),
            "quartil": value_or_none(row.iloc[16]),
            "volume_improdutivo": value_or_none(row.iloc[17]),
            "produtividade_ok_nok": value_or_none(row.iloc[18]),
            "volume_altas": value_or_none(row.iloc[19]),
            "volume_me": value_or_none(row.iloc[20]),
            "volume_mc": value_or_none(row.iloc[21]),
            "volume_reparo": value_or_none(row.iloc[22]),
            "volume_preventiva": value_or_none(row.iloc[23]),
            "volume_cleanup": value_or_none(row.iloc[24]),
            "volume_apoio": value_or_none(row.iloc[25]),
            "volume_preset": value_or_none(row.iloc[26]),
            "volume_oem": value_or_none(row.iloc[27]),
            "volume_b2b": value_or_none(row.iloc[28]),
            "volume_infra": value_or_none(row.iloc[29]),
            "sem_baixa_periodo": bool(not pd.isna(sem_baixa_value) and sem_baixa_value > 0),
            "perfil_analitico": value_or_none(extra.iloc[15]) if extra is not None else None,
            "fonte": "Analitico Nominal por Quartil + Analitico Nominal Periodo Filtrado",
        }
        output.append(record)
        technician_records.append(record)

    tech = pd.DataFrame(technician_records)
    for code, group in tech.groupby("cluster_codigo"):
        output.append(
            {
                "tipo_registro": "resumo_cluster_produtividade",
                "estado": "CE",
                "cluster_codigo": code,
                "cluster": group.iloc[0]["cluster"],
                "periodo": group.iloc[0]["periodo"],
                "tecnicos_elegiveis": int(len(group)),
                "dias_trabalhados": float(group["dias_trabalhados"].sum()),
                "volume_produtivo": float(group["volume_produtivo"].sum()),
                "produtividade_ponderada": float(
                    group["volume_produtivo"].sum() / group["dias_trabalhados"].sum()
                ),
                "atingem_meta": int(group["atingiu_meta"].sum()),
                "abaixo_meta": int((~group["atingiu_meta"]).sum()),
                "percentual_abaixo_meta": float((~group["atingiu_meta"]).mean()),
                "sem_producao": int((group["produtividade"] == 0).sum()),
                "gap_os_meta": float(group["gap_os_meta"].sum()),
                "fonte": "Analitico Nominal por Quartil",
            }
        )

    for (code, tenure), group in tech.groupby(["cluster_codigo", "faixa_tempo_casa"]):
        output.append(
            {
                "tipo_registro": "resumo_cluster_tempo_casa",
                "estado": "CE",
                "cluster_codigo": code,
                "cluster": group.iloc[0]["cluster"],
                "faixa_tempo_casa": tenure,
                "periodo": group.iloc[0]["periodo"],
                "tecnicos_elegiveis": int(len(group)),
                "meta_produtividade": float(group.iloc[0]["meta_produtividade"]),
                "produtividade_ponderada": float(
                    group["volume_produtivo"].sum() / group["dias_trabalhados"].sum()
                ) if group["dias_trabalhados"].sum() else None,
                "atingem_meta": int(group["atingiu_meta"].sum()),
                "abaixo_meta": int((~group["atingiu_meta"]).sum()),
                "sem_producao": int((group["produtividade"] == 0).sum()),
                "gap_os_meta": float(group["gap_os_meta"].sum()),
                "fonte": "Analitico Nominal por Quartil",
            }
        )

    for quartile, group in tech.groupby("quartil"):
        output.append(
            {
                "tipo_registro": "resumo_quartil_ceara",
                "estado": "CE",
                "quartil": quartile,
                "tecnicos": int(len(group)),
                "share": float(len(group) / len(tech)),
                "abaixo_meta": int((~group["atingiu_meta"]).sum()),
                "gap_os_meta": float(group["gap_os_meta"].sum()),
                "fonte": "Share por Quartil + Analitico Nominal por Quartil",
            }
        )

    detail.columns = [
        "data", "funcid", "dia_produtivo", "dia_trabalhado", "ferias",
        "admissao", "demissao", "preset", "apoio", "notdone_b2c", "ok_b2b",
        "ok_oem", "ok_infra", "outros_segmentos", "deslocamento", "execucao",
        "pendente", "suspenso", "cancelado",
    ]
    for column in detail.columns[2:]:
        detail[column] = pd.to_numeric(detail[column], errors="coerce").fillna(0)
    binary = ["dia_produtivo", "dia_trabalhado", "ferias", "admissao", "demissao"]
    volumes = list(detail.columns[7:])
    technician_day = detail.groupby(["data", "funcid"], as_index=False).agg(
        {**{column: "max" for column in binary}, **{column: "sum" for column in volumes}}
    )
    tech_lookup = tech.set_index("funcid")
    for _, row in technician_day.iterrows():
        lookup = tech_lookup.loc[row["funcid"]] if row["funcid"] in tech_lookup.index else None
        record = {
            "tipo_registro": "tecnico_dia",
            "estado": "CE",
            "data_referencia": value_or_none(row["data"]),
            "funcid": value_or_none(row["funcid"]),
            "cluster_codigo": lookup["cluster_codigo"] if lookup is not None else None,
            "cluster": lookup["cluster"] if lookup is not None else None,
            "quartil": lookup["quartil"] if lookup is not None else None,
            "dia_produtivo": bool(row["dia_produtivo"]),
            "dia_trabalhado": bool(row["dia_trabalhado"]),
            "ferias": bool(row["ferias"]),
            "admissao": bool(row["admissao"]),
            "demissao": bool(row["demissao"]),
            "volume_preset": float(row["preset"]),
            "volume_apoio": float(row["apoio"]),
            "volume_notdone_b2c": float(row["notdone_b2c"]),
            "volume_b2b": float(row["ok_b2b"]),
            "volume_oem": float(row["ok_oem"]),
            "volume_infra": float(row["ok_infra"]),
            "volume_outros_segmentos": float(row["outros_segmentos"]),
            "volume_deslocamento": float(row["deslocamento"]),
            "volume_execucao": float(row["execucao"]),
            "volume_pendente": float(row["pendente"]),
            "volume_suspenso": float(row["suspenso"]),
            "volume_cancelado": float(row["cancelado"]),
            "fonte": "Analitico Acompanhamento Tecnicos",
        }
        output.append(record)

    for _, row in daily.iterrows():
        output.append(
            {
                "tipo_registro": "acompanhamento_ceara_dia",
                "estado": "CE",
                "data_referencia": value_or_none(row.iloc[0]),
                "necessidade_folha": value_or_none(row.iloc[1]),
                "necessidade_produtiva": value_or_none(row.iloc[2]),
                "folha": value_or_none(row.iloc[3]),
                "dia_trabalhado": value_or_none(row.iloc[4]),
                "dia_produtivo": value_or_none(row.iloc[5]),
                "em_apoio": value_or_none(row.iloc[6]),
                "improdutivos": value_or_none(row.iloc[7]),
                "em_aberto": value_or_none(row.iloc[8]),
                "pendente_cancelado_suspenso": value_or_none(row.iloc[9]),
                "outros_segmentos": value_or_none(row.iloc[10]),
                "sem_baixa_dia": value_or_none(row.iloc[11]),
                "exclusivo_preset": value_or_none(row.iloc[12]),
                "ferias": value_or_none(row.iloc[13]),
                "folga_escala": value_or_none(row.iloc[14]),
                "sem_baixa_mes": value_or_none(row.iloc[15]),
                "admitidos_dia": value_or_none(row.iloc[16]),
                "demitidos_dia": value_or_none(row.iloc[17]),
                "ating_real_necessidade_folha": value_or_none(row.iloc[18]),
                "ating_real_necessidade_produtiva": value_or_none(row.iloc[19]),
                "ating_dia_trabalhado_folha": value_or_none(row.iloc[20]),
                "ating_produtivo_dia_trabalhado": value_or_none(row.iloc[21]),
                "ating_produtivo_folha": value_or_none(row.iloc[22]),
                "fonte": "Acompanhamento Tecnicos Segmento B2C",
            }
        )
    return output


def ensure_link(client: SupabaseRest, demand_id: str, base_id: str):
    current = client.request(
        "GET",
        f"/rest/v1/demanda_bases?select=base_id&demanda_id=eq.{demand_id}&base_id=eq.{base_id}",
    ) or []
    if not current:
        client.request(
            "POST", "/rest/v1/demanda_bases",
            {"demanda_id": demand_id, "base_id": base_id},
            prefer="return=minimal",
        )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", action="append", required=True, type=Path)
    parser.add_argument("--env", default=Path(".env"), type=Path)
    parser.add_argument("--project", default="3Ps")
    args = parser.parse_args()
    if len(args.file) != 5:
        raise ValueError("Informe exatamente os cinco arquivos complementares.")

    paths_by_role = {classify(path): path for path in args.file}
    expected = {
        "nominal_quartil", "share_quartil", "detalhe_diario",
        "nominal_periodo", "resumo_diario",
    }
    if set(paths_by_role) != expected:
        raise ValueError(f"Fontes incompletas: esperado {expected}, recebido {set(paths_by_role)}")

    env = parse_env(args.env)
    client = SupabaseRest(env["SUPABASE_URL"], env["SUPABASE_ANON_KEY"])
    project_filter = urllib.parse.quote(args.project, safe="")
    demands = client.request(
        "GET", f"/rest/v1/demandas?select=id&nome=eq.{project_filter}"
    ) or []
    if not demands:
        raise RuntimeError("Demanda 3Ps nao encontrada. Importe primeiro a fonte principal.")
    demand_id = demands[0]["id"]

    existing = client.request(
        "GET", "/rest/v1/bases?select=id,nome_arquivo,descricao,caminho_storage"
    ) or []
    raw_base_ids = []
    hashes: dict[str, str] = {}
    raw_counts: dict[str, int] = {}

    for role, path in paths_by_role.items():
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        hashes[role] = digest
        tag = f"sha256:{digest}"
        match = next((item for item in existing if tag in (item.get("descricao") or "")), None)
        if match:
            raw_base_ids.append(match["id"])
            ensure_link(client, demand_id, match["id"])
            continue

        formulas = load_workbook(path, data_only=False, read_only=False)
        values = load_workbook(path, data_only=True, read_only=False)
        records = raw_rows(formulas, values)
        raw_counts[role] = len(records)
        storage_path = client.upload(
            "megabrain-bases",
            f"indicadores/3ps/{digest[:12]}_{safe_name(path)}",
            path,
        )
        headers = [json_value(cell.value) for cell in values.worksheets[0][1] if cell.value is not None]
        base = create_base(
            client,
            {
                "nome_arquivo": path.name,
                "tipo_base": "outros",
                "descricao": f"Projeto 3Ps | fonte complementar {role} | {tag}",
                "qtd_linhas": len(records),
                "qtd_colunas": max(sheet.max_column for sheet in formulas.worksheets),
                "tamanho_bytes": path.stat().st_size,
                "colunas_originais": headers,
                "colunas_normalizadas": [normalize_key(item) for item in headers],
                "guardar_arquivo_original": True,
                "caminho_storage": storage_path,
                "status": "importada",
            },
        )
        rows = [
            {
                "base_id": base["id"],
                "tipo_base": "outros",
                "linha_numero": index,
                "dados": record,
            }
            for index, record in enumerate(records, 1)
        ]
        post_batches(client, "base_linhas", rows, batch_size=100)
        ensure_link(client, demand_id, base["id"])
        raw_base_ids.append(base["id"])

    combined_hash = hashlib.sha256(
        "|".join(sorted(hashes.values())).encode("utf-8")
    ).hexdigest()
    combined_tag = f"conjunto_sha256:{combined_hash}"
    curated_existing = next(
        (item for item in existing if combined_tag in (item.get("descricao") or "")), None
    )
    curated_count = 0
    if curated_existing:
        curated_base_id = curated_existing["id"]
        ensure_link(client, demand_id, curated_base_id)
    else:
        curated = build_curated(paths_by_role)
        curated_count = len(curated)
        columns = sorted({key for row in curated for key in row})
        curated_base = create_base(
            client,
            {
                "nome_arquivo": "3Ps_Ceara_fontes_complementares_2026-07-01_a_03",
                "tipo_base": "indicadores",
                "descricao": (
                    f"Projeto 3Ps | camada analitica das cinco fontes complementares | {combined_tag}. "
                    "Universo Ceara B2C, periodo de 01/07/2026 a 03/07/2026."
                ),
                "qtd_linhas": len(curated),
                "qtd_colunas": len(columns),
                "colunas_originais": columns,
                "colunas_normalizadas": columns,
                "guardar_arquivo_original": False,
                "status": "importada",
            },
        )
        curated_base_id = curated_base["id"]
        rows = [
            {
                "base_id": curated_base_id,
                "tipo_base": "indicadores",
                "linha_numero": index,
                "dados": record,
            }
            for index, record in enumerate(curated, 1)
        ]
        post_batches(client, "base_linhas", rows, batch_size=100)
        ensure_link(client, demand_id, curated_base_id)

        analysis = {
            "demanda_id": demand_id,
            "titulo": "Diagnostico complementar de produtividade do Ceara — 01 a 03/07/2026",
            "pergunta": "Onde se concentra o desvio de Producao e qual publico deve ser priorizado?",
            "resumo": (
                "Dos 127 tecnicos elegiveis, 86 (67,7%) ficaram abaixo da meta individual por tempo de casa. "
                "Fortaleza concentra 61 casos abaixo da meta e um gap estimado de 162 OS no periodo, ante "
                "25 casos e 83,5 OS no Interior."
            ),
            "evidencias": (
                "Fortaleza teve produtividade ponderada de 3,24 OS/dia e 61 de 81 tecnicos abaixo da meta. "
                "No Interior, a produtividade ponderada foi 3,63 e 25 de 46 ficaram abaixo. O maior bloco de "
                "desvio esta nos tecnicos acima de 90 dias: 52 casos em Fortaleza e 24 no Interior. Os quartis "
                "Q3, Q4 e NP somam 29 tecnicos; ha ainda tecnicos Q2 abaixo da meta absoluta, portanto quartil "
                "nao deve substituir a regua de meta por tempo de casa."
            ),
            "hipoteses": (
                "A conversao de presenca em dia produtivo oscilou entre 90,0% e 93,6%. As bases registram apoio, "
                "outros segmentos e tecnicos sem baixa, mas a causalidade individual deve ser validada antes do FCA."
            ),
            "sugestoes": (
                "Priorizar Fortaleza e os tecnicos acima de 90 dias; abrir plano nominal para os 86 abaixo da meta, "
                "com tratamento imediato dos oito NP e acompanhamento diario de dia trabalhado versus produtivo."
            ),
            "proximos_passos": (
                "Cruzar ausencias, escala, ordens, ferramental e lideranca por FUNCID; validar as causas com GOs e GAs; "
                "converter os grupos priorizados em FCA com responsavel e prazo."
            ),
        }
        client.request("POST", "/rest/v1/analises", analysis, prefer="return=minimal")

    client.request(
        "POST", "/rest/v1/logs",
        {
            "demanda_id": demand_id,
            "base_id": curated_base_id,
            "tipo": "upload",
            "mensagem": "Importacao das cinco fontes complementares de produtividade do projeto 3Ps.",
            "detalhes": {
                "hashes": hashes,
                "bases_brutas": raw_base_ids,
                "base_analitica": curated_base_id,
                "linhas_analiticas_novas": curated_count,
            },
        },
        prefer="return=minimal",
    )

    print(json.dumps({
        "status": "importado" if curated_count else "ja_importado",
        "demanda_id": demand_id,
        "bases_brutas": raw_base_ids,
        "base_analitica": curated_base_id,
        "linhas_brutas_novas": raw_counts,
        "linhas_analiticas_novas": curated_count,
        "hashes": hashes,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
