"""Importa o terceiro lote de produtividade e reconcilia duplicidades do 3Ps."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import urllib.parse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from importar_workbook_3ps import (
    SupabaseRest,
    create_base,
    json_value,
    normalize_key,
    parse_env,
    post_batches,
    raw_rows,
)


def classify(path: Path) -> str:
    name = normalize_key(path.stem)
    if "baremo_ok" in name:
        return "nominal_baremo_ok"
    if "qtd_h_c" in name:
        return "headcount"
    if "ok_e_nok" in name:
        return "produtividade_ok_nok"
    if name.endswith("ok"):
        return "produtividade_ok"
    raise ValueError(f"Fonte nao reconhecida: {path.name}")


def safe_name(path: Path) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", path.name)


def scalar(value):
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        value = value.item()
    return json_value(value)


def cluster_fields(value) -> tuple[str | None, str]:
    text = str(value)
    if "C.26" in text:
        return "C.26", "Ceara Interior"
    if "C.27" in text:
        return "C.27", "Fortaleza"
    if text == "Total":
        return "CE", "Ceara"
    return None, text


def aggregate_series(paths: dict[str, Path]) -> list[dict]:
    output: list[dict] = []
    metric_names = {
        "headcount": ("headcount_dia_trabalhado", "tecnicos"),
        "produtividade_ok": ("produtividade_ok", "OS OK/dia"),
        "produtividade_ok_nok": ("produtividade_ok_nok", "OS/dia"),
    }
    monthly = [(2, "FEV"), (3, "MAR"), (4, "ABR"), (5, "MAI"), (6, "JUN"), (7, "JUL")]
    weekly = [(10, "S25"), (11, "S26"), (12, "S27"), (13, "S28")]
    for role in ("headcount", "produtividade_ok_nok", "produtividade_ok"):
        frame = pd.read_excel(paths[role], header=None)
        metric, unit = metric_names[role]
        # Linhas 26 a 28 no Excel: Total CE, C.26 e C.27.
        for excel_row in (26, 27, 28):
            row = frame.iloc[excel_row - 1]
            code, cluster = cluster_fields(row.iloc[1])
            for column, period in monthly + weekly:
                value = row.iloc[column]
                if pd.isna(value):
                    continue
                period_type = "mes" if column <= 7 else "semana"
                record = {
                    "tipo_registro": "serie_bi_produtividade",
                    "estado": "CE",
                    "cluster_codigo": code,
                    "cluster": cluster,
                    "metrica": metric,
                    "unidade": unit,
                    "tipo_periodo": period_type,
                    "periodo": period,
                    "valor": scalar(value),
                    "fonte": paths[role].name,
                    "tratamento_duplicidade": "atualiza_nao_soma",
                    "atualizacao_fonte": "2026-07-06",
                }
                if period == "S27" and role == "produtividade_ok" and code in ("C.26", "C.27"):
                    record["duplica_indicador_principal"] = True
                    record["fonte_preferencial"] = True
                    record["motivo_precedencia"] = "Exportacao BI posterior ao fechamento da S27"
                if period == "S28":
                    record["periodo_parcial"] = True
                output.append(record)
    return output


def baremo_records(path: Path, existing_ids: set[str]) -> list[dict]:
    frame = pd.read_excel(path)
    frame = frame[frame.iloc[:, 9].notna()].copy()
    output = []
    for _, row in frame.iterrows():
        code, cluster = cluster_fields(row.iloc[3])
        funcid = str(row.iloc[9])
        output.append(
            {
                "tipo_registro": "tecnico_produtividade_baremo_ok",
                "estado": "CE",
                "cluster_codigo": code,
                "cluster": cluster,
                "periodo": scalar(row.iloc[0]),
                "semana": scalar(row.iloc[1]),
                "gerente_geral": scalar(row.iloc[4]),
                "gestor_operacoes": scalar(row.iloc[5]),
                "gestor_area": scalar(row.iloc[6]),
                "tecnico": scalar(row.iloc[7]),
                "cargo": scalar(row.iloc[8]),
                "funcid": funcid,
                "admissao": scalar(row.iloc[10]),
                "faixa_tempo_casa": scalar(row.iloc[11]),
                "volume_produtivo_baremo": scalar(row.iloc[12]),
                "dias_trabalhados": scalar(row.iloc[13]),
                "produtividade_baremo": scalar(row.iloc[14]),
                "quartil_baremo": scalar(row.iloc[15]),
                "volume_improdutivo": scalar(row.iloc[16]),
                "produtividade_ok_nok_baremo": scalar(row.iloc[17]),
                "volume_altas_baremo": scalar(row.iloc[18]),
                "volume_me_baremo": scalar(row.iloc[19]),
                "volume_mc": scalar(row.iloc[20]),
                "volume_reparo": scalar(row.iloc[21]),
                "volume_preventiva": scalar(row.iloc[22]),
                "volume_upgrade_baremo": scalar(row.iloc[23]),
                "volume_cleanup": scalar(row.iloc[24]),
                "volume_apoio": scalar(row.iloc[25]),
                "volume_preset": scalar(row.iloc[26]),
                "volume_oem": scalar(row.iloc[27]),
                "volume_b2b": scalar(row.iloc[28]),
                "volume_infra": scalar(row.iloc[29]),
                "sobrepoe_nominal_11_ok": funcid in existing_ids,
                "tratamento_duplicidade": "metrica_alternativa_nao_somar",
                "fonte": path.name,
            }
        )
    return output


def screenshot_records() -> list[dict]:
    common = {
        "estado": "CE",
        "periodo": "01/07/2026 a 03/07/2026",
        "segmento": "B2C",
        "tipo_atividade": "REPARO",
        "tempo_treinamento": "Tempo de Treinamento ultrapassado",
        "atualizacao_painel": "2026-07-06T13:16:08",
        "fonte": "Print Acompanhamento de Indicadores de Produtividade",
    }
    output = [
        {
            **common,
            "tipo_registro": "evidencia_painel_filtros",
            "observacao": "Transcricao do print fornecido; a imagem nao estava disponivel como arquivo local.",
        }
    ]
    technicians = [
        ("ITALO DE MOURA BERNARDINO", 10),
        ("ANDRE JEFFERSON SILVA DE SOUZA", 7),
        ("JADSON ARAUJO FERREIRA", 7),
        ("LUAN RICARDO DE SOUZA LIMA", 7),
        ("ADRIANO SANTOS DE SOUSA", 6),
        ("FRANCISCO GIRLENO DOS SANTOS DE OLIVEIRA", 6),
        ("LUIZ NAUA LIMA DA SILVA", 6),
        ("REGIVANIO DOS SANTOS LIRA", 6),
        ("THIAGO NERI URCULINO", 6),
    ]
    areas = [
        ("PEDRO HENRIQUE MARQUES FERREIRA", 82),
        ("THIAGO LEITE RABELO", 47),
        ("DERLANDIO CORREIA DO NASCIMENTO", 37),
        ("CICERO NATANAEL DE SOUZA FREIRE", 31),
        ("JOSE ORLANDO BARBOSA", 31),
        ("LEONARDO XAVIER CARTAXO", 27),
        ("CLERISTON ARAUJO BENEVIDES MACHADO", 26),
        ("AGENOR VIEIRA DOS SANTOS JUNIOR", 20),
        ("JOSE GERALDO BARBOSA DA SILVA", 14),
    ]
    operations = [
        ("JOSE NILTON MENDES LIMA", 213),
        ("JOSE GERALDO BARBOSA DA SILVA", 95),
        ("JEFFERSON OLIVEIRA DA SILVA", 20),
        ("FRANCISCO CLEITON CRUZ DO NASCIMENTO", 0),
    ]
    for level, items in (("tecnico", technicians), ("GA", areas), ("GO", operations)):
        for rank, (name, value) in enumerate(items, 1):
            output.append(
                {
                    **common,
                    "tipo_registro": "ranking_improdutividade_painel",
                    "nivel": level,
                    "ranking": rank,
                    "nome": name,
                    "volume_improdutivo": value,
                    "evidencia_status": "validacao_duplicada",
                    "tratamento_duplicidade": "nao_somar",
                }
            )
    motives = [
        ("Abertura indevida", 59), ("Cliente ausente", 52),
        ("Solicitacao de reagendamento", 36), ("Falha massiva", 23),
        ("Desistiu do servico", 9), ("Area de risco", 5), ("Triagem CST", 5),
        ("Troca conector externo", 5), ("Visita a campo evitada", 5),
        ("Drop refeito", 4), ("Entrada nao autorizada", 3),
        ("Troca conector interno", 2), ("Troca ONU", 2), ("Chuva", 1),
        ("Falta material", 1),
    ]
    for rank, (name, value) in enumerate(motives, 1):
        output.append(
            {
                **common,
                "tipo_registro": "motivo_baixa_improdutiva",
                "ranking": rank,
                "motivo": name,
                "quantidade": value,
                "evidencia_status": "novo_no_conjunto",
                "recorte_visivel_completo": True,
            }
        )
    submotives = [
        ("Casa fechada", 41), ("Reagendamento", 36),
        ("Demanda nao e tecnica", 25), ("Cliente nao solicitou servico", 17),
        ("Sinal e potencia normalizados", 17), ("Sem potencia", 13),
        ("Cliente nao estava em casa", 11), ("Potencia fora do padrao", 7),
        ("Servico nao e mais necessario", 7), ("Desgaste natural", 5),
        ("TRIADOCST_Online", 5), ("Risco a integridade", 4),
        ("Servico nao autorizado", 3), ("Em branco", 2),
        ("Acoes de terceiros", 2), ("Carga alta", 2),
        ("FTTA com problema", 2), ("Inclusao em Massiva pos Abertura OS", 2),
        ("Instrucoes ao cliente foram suficientes", 2),
    ]
    for rank, (name, value) in enumerate(submotives, 1):
        output.append(
            {
                **common,
                "tipo_registro": "submotivo_baixa_improdutiva",
                "ranking": rank,
                "submotivo": name,
                "quantidade": value,
                "evidencia_status": "novo_no_conjunto",
                "recorte_visivel_incompleto": True,
            }
        )
    output.append(
        {
            **common,
            "tipo_registro": "resumo_motivos_improdutividade",
            "volume_improdutivo_validado": 328,
            "baixas_com_motivo_visivel": 212,
            "top_4_motivos_quantidade": 170,
            "top_4_motivos_share_das_baixas_classificadas": 170 / 212,
            "top_4_motivos": "Abertura indevida; Cliente ausente; Solicitacao de reagendamento; Falha massiva",
            "observacao": "O ranking de 328 replica o volume nominal; os 212 motivos nao devem ser extrapolados para todo o volume sem validar a cobertura.",
        }
    )
    return output


def ensure_link(client: SupabaseRest, demand_id: str, base_id: str):
    current = client.request(
        "GET",
        f"/rest/v1/demanda_bases?select=base_id&demanda_id=eq.{demand_id}&base_id=eq.{base_id}",
    ) or []
    if not current:
        client.request(
            "POST", "/rest/v1/demanda_bases",
            {"demanda_id": demand_id, "base_id": base_id},
            prefer="return=minimal",
        )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", action="append", required=True, type=Path)
    parser.add_argument("--env", default=Path(".env"), type=Path)
    parser.add_argument("--project", default="3Ps")
    args = parser.parse_args()
    if len(args.file) != 4:
        raise ValueError("Informe exatamente os quatro arquivos deste lote.")
    paths = {classify(path): path for path in args.file}
    expected = {"headcount", "produtividade_ok_nok", "produtividade_ok", "nominal_baremo_ok"}
    if set(paths) != expected:
        raise ValueError(f"Fontes incompletas: {set(paths)}")

    env = parse_env(args.env)
    client = SupabaseRest(env["SUPABASE_URL"], env["SUPABASE_ANON_KEY"])
    project_filter = urllib.parse.quote(args.project, safe="")
    demand = client.request(
        "GET", f"/rest/v1/demandas?select=id&nome=eq.{project_filter}"
    )[0]
    demand_id = demand["id"]
    existing = client.request(
        "GET", "/rest/v1/bases?select=id,nome_arquivo,descricao,caminho_storage"
    ) or []

    raw_ids = []
    hashes = {}
    for role, path in paths.items():
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        hashes[role] = digest
        tag = f"sha256:{digest}"
        match = next((item for item in existing if tag in (item.get("descricao") or "")), None)
        if match:
            raw_ids.append(match["id"])
            ensure_link(client, demand_id, match["id"])
            continue
        formulas = load_workbook(path, data_only=False, read_only=False)
        values = load_workbook(path, data_only=True, read_only=False)
        records = raw_rows(formulas, values)
        storage_path = client.upload(
            "megabrain-bases", f"indicadores/3ps/{digest[:12]}_{safe_name(path)}", path
        )
        headers = [json_value(cell.value) for cell in values.worksheets[0][1] if cell.value is not None]
        base = create_base(
            client,
            {
                "nome_arquivo": path.name,
                "tipo_base": "outros",
                "descricao": f"Projeto 3Ps | lote produtividade {role} | {tag}",
                "qtd_linhas": len(records),
                "qtd_colunas": max(sheet.max_column for sheet in formulas.worksheets),
                "tamanho_bytes": path.stat().st_size,
                "colunas_originais": headers,
                "colunas_normalizadas": [normalize_key(item) for item in headers],
                "guardar_arquivo_original": True,
                "caminho_storage": storage_path,
                "status": "importada",
            },
        )
        post_batches(
            client,
            "base_linhas",
            [
                {"base_id": base["id"], "tipo_base": "outros", "linha_numero": index, "dados": row}
                for index, row in enumerate(records, 1)
            ],
            batch_size=100,
        )
        ensure_link(client, demand_id, base["id"])
        raw_ids.append(base["id"])

    combined_hash = hashlib.sha256("|".join(sorted(hashes.values())).encode()).hexdigest()
    combined_tag = f"conjunto_sha256:{combined_hash}"
    curated_existing = next(
        (item for item in existing if combined_tag in (item.get("descricao") or "")), None
    )
    curated_count = 0
    if curated_existing:
        curated_id = curated_existing["id"]
        ensure_link(client, demand_id, curated_id)
    else:
        prior_rows = client.request(
            "GET",
            "/rest/v1/base_linhas?base_id=eq.10007154-8266-4eb7-aae3-bd095040cd03&select=dados&limit=1000",
        ) or []
        existing_ids = {
            str(item["dados"].get("funcid")) for item in prior_rows
            if item.get("dados", {}).get("tipo_registro") == "tecnico_produtividade_periodo"
        }
        curated = aggregate_series(paths) + baremo_records(paths["nominal_baremo_ok"], existing_ids) + screenshot_records()
        curated_count = len(curated)
        columns = sorted({key for row in curated for key in row})
        base = create_base(
            client,
            {
                "nome_arquivo": "3Ps_Ceara_produtividade_reconciliada_lote_3",
                "tipo_base": "indicadores",
                "descricao": (
                    f"Projeto 3Ps | series BI, baremo OK e evidencia do painel de improdutividade | {combined_tag}. "
                    "Duplicidades marcadas para atualizacao, validacao ou metrica alternativa; nunca para soma."
                ),
                "qtd_linhas": len(curated),
                "qtd_colunas": len(columns),
                "colunas_originais": columns,
                "colunas_normalizadas": columns,
                "guardar_arquivo_original": False,
                "status": "importada",
            },
        )
        curated_id = base["id"]
        post_batches(
            client,
            "base_linhas",
            [
                {"base_id": curated_id, "tipo_base": "indicadores", "linha_numero": index, "dados": row}
                for index, row in enumerate(curated, 1)
            ],
            batch_size=100,
        )
        ensure_link(client, demand_id, curated_id)
        client.request(
            "POST", "/rest/v1/analises",
            {
                "demanda_id": demand_id,
                "titulo": "Reconciliacao da produtividade e motivos de improdutividade",
                "pergunta": "O que e informacao nova e o que apenas replica as fontes anteriores?",
                "resumo": (
                    "Os rankings do print replicam exatamente as 328 improdutivas do nominal ja importado e nao foram somados. "
                    "A exportacao BI posterior atualiza a S27 para 3,5 OS OK/dia no Interior e 3,2 em Fortaleza. "
                    "O baremo OK e uma metrica alternativa, nao aditiva, e eleva em media 0,27 OS/dia os 127 tecnicos comparaveis."
                ),
                "evidencias": (
                    "O volume improdutivo divide-se em 213 sob Jose Nilton, 95 sob Jose Geraldo e 20 sob Jefferson, total 328. "
                    "Entre 212 baixas com motivo visivel, abertura indevida, cliente ausente, reagendamento e falha massiva somam 170 (80,2%). "
                    "No baremo, 18 dos 127 tecnicos comparaveis mudam para um quartil superior; oito tecnicos adicionais aparecem no universo de 135."
                ),
                "hipoteses": (
                    "A improdutividade tem componente relevante de qualidade da demanda e disponibilidade do cliente, alem da execucao de campo. "
                    "A cobertura de motivos e menor que o volume total de improdutivas e precisa ser validada antes de generalizar a causa."
                ),
                "sugestoes": (
                    "Usar Produtividade OK como indicador executivo oficial, manter Baremo OK apenas como lente complementar e atacar os quatro motivos "
                    "mais frequentes por origem da demanda, confirmacao com cliente e tratamento de massivas."
                ),
                "proximos_passos": (
                    "Validar a cobertura dos 212 motivos sobre as 328 improdutivas; separar causas controlaveis pela Operacao das externas; "
                    "definir dono, prazo e meta de reducao para cada causa no FCA."
                ),
            },
            prefer="return=minimal",
        )

    client.request(
        "POST", "/rest/v1/logs",
        {
            "demanda_id": demand_id,
            "base_id": curated_id,
            "tipo": "upload",
            "mensagem": "Importacao e reconciliacao do terceiro lote de produtividade do projeto 3Ps.",
            "detalhes": {
                "hashes": hashes,
                "bases_brutas": raw_ids,
                "base_analitica": curated_id,
                "linhas_analiticas_novas": curated_count,
                "regra_duplicidade": "atualiza_nao_soma",
            },
        },
        prefer="return=minimal",
    )
    print(json.dumps({
        "status": "importado" if curated_count else "ja_importado",
        "demanda_id": demand_id,
        "bases_brutas": raw_ids,
        "base_analitica": curated_id,
        "linhas_analiticas_novas": curated_count,
        "hashes": hashes,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
