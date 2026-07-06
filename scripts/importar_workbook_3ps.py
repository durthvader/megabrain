"""Importa o workbook principal do projeto 3Ps no Supabase do Megabrain.

Cria uma base bruta, preservando valores calculados e formulas de todas as abas,
e uma base analitica normalizada restrita aos clusters C.26 e C.27 do Ceara.
O script e idempotente por SHA-256 do arquivo de origem.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import mimetypes
import re
import secrets
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def json_value(value):
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def normalize_key(value: object) -> str:
    text = str(value or "campo")
    text = "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    ).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "campo"


class SupabaseRest:
    def __init__(self, url: str, key: str):
        self.url = url.rstrip("/")
        self.key = key

    def request(self, method: str, path: str, payload=None, prefer: str | None = None):
        url = f"{self.url}{path}"
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
        }
        data = None
        if payload is not None:
            data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            headers["Content-Type"] = "application/json; charset=utf-8"
        if prefer:
            headers["Prefer"] = prefer
        request = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                body = response.read()
                return json.loads(body.decode("utf-8")) if body else None
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Supabase {method} {path}: HTTP {exc.code}: {detail}") from exc

    def upload(self, bucket: str, object_path: str, file_path: Path) -> str:
        quoted = urllib.parse.quote(object_path, safe="/")
        url = f"{self.url}/storage/v1/object/{bucket}/{quoted}"
        headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": mimetypes.guess_type(file_path.name)[0]
            or "application/octet-stream",
            "x-upsert": "true",
        }
        request = urllib.request.Request(
            url, data=file_path.read_bytes(), headers=headers, method="POST"
        )
        try:
            with urllib.request.urlopen(request, timeout=120):
                return object_path
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Storage upload: HTTP {exc.code}: {detail}") from exc


def workbook_manifest(workbook) -> list[dict]:
    manifest = []
    for sheet in workbook.worksheets:
        manifest.append(
            {
                "aba": sheet.title,
                "estado": sheet.sheet_state,
                "dimensao": sheet.calculate_dimension(),
                "linhas": sheet.max_row,
                "colunas": sheet.max_column,
                "mesclagens": [str(item) for item in sheet.merged_cells.ranges],
            }
        )
    return manifest


def raw_rows(workbook_formulas, workbook_values) -> list[dict]:
    rows = [
        {
            "aba": "_manifesto",
            "linha_excel": 0,
            "workbook": workbook_manifest(workbook_formulas),
        }
    ]
    for sheet_formula in workbook_formulas.worksheets:
        sheet_values = workbook_values[sheet_formula.title]
        for row_number in range(1, sheet_formula.max_row + 1):
            values: dict[str, object] = {}
            formulas: dict[str, str] = {}
            for column_number in range(1, sheet_formula.max_column + 1):
                key = get_column_letter(column_number).lower()
                formula_value = sheet_formula.cell(row_number, column_number).value
                cached_value = sheet_values.cell(row_number, column_number).value
                if formula_value is None and cached_value is None:
                    continue
                values[key] = json_value(cached_value)
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    formulas[key] = formula_value
                elif cached_value is None:
                    values[key] = json_value(formula_value)
            if values or formulas:
                record = {
                    "aba": sheet_formula.title,
                    "linha_excel": row_number,
                    "valores": values,
                }
                if formulas:
                    record["formulas"] = formulas
                rows.append(record)
    return rows


def add_metric(
    output: list[dict], *, cluster_code: str, cluster: str, indicator: str,
    period: str, value, target, unit: str, source: str,
    tenure: str | None = None, date_start: str | None = None,
    date_end: str | None = None, date_reference: str | None = None,
):
    numeric_value = value if isinstance(value, (int, float)) else None
    status = "sem_dado" if numeric_value is None else (
        "atingida" if numeric_value >= target else "abaixo"
    )
    output.append(
        {
            "tipo_registro": "indicador_executivo",
            "estado": "CE",
            "cluster_codigo": cluster_code,
            "cluster": cluster,
            "indicador": indicator,
            "faixa_tempo_casa": tenure,
            "periodo": period,
            "data_inicio": date_start,
            "data_fim": date_end,
            "data_referencia": date_reference,
            "realizado": numeric_value,
            "meta": target,
            "gap": None if numeric_value is None else numeric_value - target,
            "unidade": unit,
            "status_meta": status,
            "fonte": source,
        }
    )


def curated_rows(workbook_values) -> list[dict]:
    output: list[dict] = []
    dashboard = workbook_values["painel"]
    cluster_rows = {
        24: ("C.26", "Ceara Interior"),
        25: ("C.27", "Fortaleza"),
    }

    periods = {
        "S25": ("2026-06-15", "2026-06-21"),
        "S26": ("2026-06-22", "2026-06-28"),
        "S27": ("2026-06-29", "2026-07-05"),
    }
    for row, (code, cluster) in cluster_rows.items():
        add_metric(
            output, cluster_code=code, cluster=cluster, indicator="Presenca",
            period="03/07/2026", value=dashboard.cell(row, 4).value,
            target=0.95, unit="percentual", source=f"painel!D{row}",
            date_reference="2026-07-03",
        )
        production_blocks = [
            ("Ate 59 dias", 3.0, (5, 6, 7)),
            ("60 a 89 dias", 3.5, (8, 9, 10)),
            ("Acima de 90 dias", 4.0, (11, 12, 13)),
        ]
        for tenure, target, columns in production_blocks:
            for period, column in zip(("S25", "S26", "S27"), columns):
                start, end = periods[period]
                add_metric(
                    output, cluster_code=code, cluster=cluster,
                    indicator="Producao", tenure=tenure, period=period,
                    value=dashboard.cell(row, column).value, target=target,
                    unit="OS OK/dia", source=f"painel!{get_column_letter(column)}{row}",
                    date_start=start, date_end=end,
                )
        for period, column in zip(("S25", "S26", "S27"), (14, 15, 16)):
            start, end = periods[period]
            add_metric(
                output, cluster_code=code, cluster=cluster, indicator="Prazo 24h",
                period=period, value=dashboard.cell(row, column).value,
                target=0.85, unit="percentual",
                source=f"painel!{get_column_letter(column)}{row}",
                date_start=start, date_end=end,
            )

    presence = workbook_values.worksheets[6]
    for row, (code, cluster) in {15: ("C.26", "Ceara Interior"), 16: ("C.27", "Fortaleza")}.items():
        planned = presence.cell(row, 2).value
        for column in range(3, 36):
            date_value = presence.cell(1, column).value
            actual = presence.cell(row, column).value
            if not isinstance(date_value, (dt.date, dt.datetime)) or actual is None:
                continue
            day_label = str(presence.cell(2, column).value or "")
            output.append(
                {
                    "tipo_registro": "presenca_diaria",
                    "estado": "CE",
                    "cluster_codigo": code,
                    "cluster": cluster,
                    "data_referencia": date_value.date().isoformat()
                    if isinstance(date_value, dt.datetime) else date_value.isoformat(),
                    "dia": day_label,
                    "presenca_planejada": planned,
                    "presenca_realizada": actual,
                    "taxa_bruta": actual / planned if planned else None,
                    "comparavel_meta": not any(
                        item in normalize_key(day_label) for item in ("sab", "dom", "feriado")
                    ),
                    "fonte": f"presenca!{get_column_letter(column)}{row}",
                }
            )

    backlog = workbook_values["hist Backlog"]
    for row, (code, cluster) in {24: ("C.26", "Ceara Interior"), 25: ("C.27", "Fortaleza")}.items():
        current_date = None
        for column in range(2, backlog.max_column + 1):
            header_date = backlog.cell(2, column).value
            if isinstance(header_date, (dt.date, dt.datetime)):
                current_date = header_date
            metric = backlog.cell(3, column).value
            value = backlog.cell(row, column).value
            if current_date is None or metric is None or value is None:
                continue
            metric_key = normalize_key(metric)
            date_text = current_date.date().isoformat() if isinstance(current_date, dt.datetime) else current_date.isoformat()
            output.append(
                {
                    "tipo_registro": "backlog_diario",
                    "estado": "CE",
                    "cluster_codigo": code,
                    "cluster": cluster,
                    "data_referencia": date_text,
                    "metrica": metric_key,
                    "valor": value,
                    "meta": 1.0 if metric_key == "dias_estoque" else None,
                    "status_meta": (
                        "atingida" if metric_key == "dias_estoque" and value <= 1.0
                        else "abaixo" if metric_key == "dias_estoque" else None
                    ),
                    "fonte": f"hist Backlog!{get_column_letter(column)}{row}",
                }
            )

    daily = workbook_values.worksheets[8]
    header_counts: dict[str, int] = {}
    headers: list[str] = []
    for column in range(1, daily.max_column + 1):
        base = normalize_key(daily.cell(1, column).value or get_column_letter(column))
        header_counts[base] = header_counts.get(base, 0) + 1
        suffix = header_counts[base]
        headers.append(base if suffix == 1 else f"{base}_{suffix}")
    for row in range(2, daily.max_row + 1):
        cluster_value = daily.cell(row, 1).value
        if cluster_value not in ("CE/Interior", "Fortaleza"):
            continue
        situation = str(daily.cell(row, 29).value or "").upper()
        technician_name = str(daily.cell(row, 6).value or "")
        if situation == "B2B" or "(B2B)" in technician_name.upper():
            continue
        record = {
            "tipo_registro": "tecnico_acompanhamento",
            "estado": "CE",
            "cluster_codigo": "C.26" if cluster_value == "CE/Interior" else "C.27",
            "cluster": "Ceara Interior" if cluster_value == "CE/Interior" else "Fortaleza",
            "linha_excel": row,
            "escopo": "B2C",
            "fonte": f"Versao Daily!A{row}:AQ{row}",
        }
        for column, header in enumerate(headers, 1):
            value = daily.cell(row, column).value
            if value is not None:
                record[header] = json_value(value)
        output.append(record)
    return output


def post_batches(client: SupabaseRest, table: str, rows: list[dict], batch_size: int = 100):
    for start in range(0, len(rows), batch_size):
        client.request(
            "POST", f"/rest/v1/{table}", rows[start:start + batch_size],
            prefer="return=minimal",
        )


def create_base(client: SupabaseRest, payload: dict) -> dict:
    result = client.request(
        "POST", "/rest/v1/bases", payload,
        prefer="return=representation",
    )
    return result[0]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, type=Path)
    parser.add_argument("--env", default=Path(".env"), type=Path)
    parser.add_argument("--project", default="3Ps")
    args = parser.parse_args()

    env = parse_env(args.env)
    client = SupabaseRest(env["SUPABASE_URL"], env["SUPABASE_ANON_KEY"])
    source_hash = hashlib.sha256(args.file.read_bytes()).hexdigest()
    hash_tag = f"sha256:{source_hash}"

    query_name = urllib.parse.quote(args.project, safe="")
    demands = client.request(
        "GET",
        f"/rest/v1/demandas?select=id,nome,tipo,status,descricao&nome=eq.{query_name}",
    ) or []
    if demands:
        demand = demands[0]
    else:
        demand = client.request(
            "POST",
            "/rest/v1/demandas",
            {
                "nome": args.project,
                "tipo": "indicadores",
                "descricao": (
                    "Gestao operacional dos 3 Ps no Ceara: Presenca, Producao e Prazo. "
                    "Escopo executivo dos clusters C.26 Ceara Interior e C.27 Fortaleza."
                ),
                "data_inicio": "2026-07-06",
                "status": "ativa",
                "token_publico": secrets.token_urlsafe(12)[:12],
            },
            prefer="return=representation",
        )[0]

    existing = client.request(
        "GET", "/rest/v1/bases?select=id,nome_arquivo,descricao,caminho_storage"
    ) or []
    existing_same_hash = [item for item in existing if hash_tag in (item.get("descricao") or "")]
    if existing_same_hash:
        base_ids = [item["id"] for item in existing_same_hash]
        current_links = client.request(
            "GET",
            f"/rest/v1/demanda_bases?select=base_id&demanda_id=eq.{demand['id']}",
        ) or []
        linked_ids = {item["base_id"] for item in current_links}
        missing_links = [
            {"demanda_id": demand["id"], "base_id": base_id}
            for base_id in base_ids if base_id not in linked_ids
        ]
        if missing_links:
            post_batches(client, "demanda_bases", missing_links, batch_size=100)
        print(json.dumps({
            "status": "ja_importado",
            "demanda_id": demand["id"],
            "base_ids": base_ids,
            "sha256": source_hash,
        }))
        return

    formulas = load_workbook(args.file, data_only=False, read_only=False)
    values = load_workbook(args.file, data_only=True, read_only=False)
    raw = raw_rows(formulas, values)
    curated = curated_rows(values)

    storage_path = None
    storage_error = None
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", args.file.name)
    object_path = f"indicadores/3ps/{source_hash[:12]}_{safe_name}"
    try:
        storage_path = client.upload("megabrain-bases", object_path, args.file)
    except RuntimeError as exc:
        storage_error = str(exc)

    raw_base = create_base(
        client,
        {
            "nome_arquivo": args.file.name,
            "tipo_base": "outros",
            "descricao": (
                f"Projeto 3Ps | fonte principal bruta | 13 abas | {hash_tag}. "
                "Valores calculados e formulas preservados por linha e aba."
            ),
            "qtd_linhas": len(raw),
            "qtd_colunas": max(sheet.max_column for sheet in formulas.worksheets),
            "tamanho_bytes": args.file.stat().st_size,
            "colunas_originais": ["aba", "linha_excel", "valores", "formulas"],
            "colunas_normalizadas": ["aba", "linha_excel", "valores", "formulas"],
            "guardar_arquivo_original": storage_path is not None,
            "caminho_storage": storage_path,
            "status": "importada",
        },
    )
    raw_db_rows = [
        {
            "base_id": raw_base["id"],
            "tipo_base": "outros",
            "linha_numero": index,
            "dados": row,
        }
        for index, row in enumerate(raw, 1)
    ]
    post_batches(client, "base_linhas", raw_db_rows, batch_size=75)

    all_curated_columns = sorted({key for row in curated for key in row})
    curated_base = create_base(
        client,
        {
            "nome_arquivo": "3Ps_Ceara_base_analitica",
            "tipo_base": "indicadores",
            "descricao": (
                f"Projeto 3Ps | base analitica somente Ceara C.26 e C.27 | {hash_tag}. "
                "Indicadores executivos, presenca diaria, backlog e acompanhamento B2C."
            ),
            "qtd_linhas": len(curated),
            "qtd_colunas": len(all_curated_columns),
            "tamanho_bytes": None,
            "colunas_originais": all_curated_columns,
            "colunas_normalizadas": all_curated_columns,
            "guardar_arquivo_original": False,
            "caminho_storage": None,
            "status": "importada",
        },
    )
    curated_db_rows = [
        {
            "base_id": curated_base["id"],
            "tipo_base": "indicadores",
            "linha_numero": index,
            "dados": row,
        }
        for index, row in enumerate(curated, 1)
    ]
    post_batches(client, "base_linhas", curated_db_rows, batch_size=100)

    links = [
        {"demanda_id": demand["id"], "base_id": raw_base["id"]},
        {"demanda_id": demand["id"], "base_id": curated_base["id"]},
    ]
    post_batches(client, "demanda_bases", links, batch_size=100)

    analysis = {
        "demanda_id": demand["id"],
        "titulo": "Leitura preliminar da base principal do Ceara",
        "pergunta": "Quais desvios dos 3 Ps exigem acao prioritaria no Ceara?",
        "resumo": (
            "Fortaleza (C.27) esta abaixo das metas de Presenca, Producao e Prazo. "
            "Ceara Interior (C.26) supera a meta de Presenca, mantem o Prazo no limite "
            "e concentra o desvio em Producao, sobretudo nos tecnicos acima de 90 dias."
        ),
        "evidencias": (
            "Em 03/07, Presenca foi 94,5% em Fortaleza e 102,5% no Interior, ante meta de 95%. "
            "Na S27, o Prazo 24h foi 83,2% em Fortaleza e 85,2% no Interior, ante meta de 85%. "
            "Na Producao S27, Fortaleza ficou abaixo da meta nas tres faixas de tempo de casa "
            "(2,4; 2,5; 3,1). No Interior, tecnicos acima de 90 dias registraram 3,3 ante meta 4,0."
        ),
        "hipoteses": (
            "As abas de acompanhamento apontam ocorrencias de operacao assistida, ferramental, "
            "estoque, treinamento e movimentacoes de RH. A relacao causal com os indicadores "
            "permanece em validacao ate a chegada das demais bases."
        ),
        "sugestoes": (
            "Priorizar C.27 no FCA dos tres Ps e abrir analise individual por faixa de tempo de casa. "
            "Para C.26, concentrar a recuperacao na produtividade dos tecnicos acima de 90 dias e "
            "monitorar o Prazo, que encerrou a S27 apenas 0,2 p.p. acima da meta."
        ),
        "proximos_passos": (
            "Cruzar as proximas bases por matricula, cluster e data; validar causas com os gestores; "
            "definir responsaveis, prazos e status para o FCA executivo."
        ),
    }
    client.request("POST", "/rest/v1/analises", analysis, prefer="return=minimal")

    log_payload = {
        "demanda_id": demand["id"],
        "base_id": raw_base["id"],
        "tipo": "upload",
        "mensagem": f"Importacao da fonte principal do projeto 3Ps: {len(raw)} linhas brutas e {len(curated)} linhas analiticas.",
        "detalhes": {
            "sha256": source_hash,
            "abas": formulas.sheetnames,
            "base_bruta_id": raw_base["id"],
            "base_analitica_id": curated_base["id"],
            "storage_path": storage_path,
            "storage_error": storage_error,
        },
    }
    client.request("POST", "/rest/v1/logs", log_payload, prefer="return=minimal")

    print(json.dumps({
        "status": "importado",
        "demanda_id": demand["id"],
        "base_bruta_id": raw_base["id"],
        "base_analitica_id": curated_base["id"],
        "linhas_brutas": len(raw),
        "linhas_analiticas": len(curated),
        "sha256": source_hash,
        "storage_path": storage_path,
        "storage_error": storage_error,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
